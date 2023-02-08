from flask import Flask, render_template, request, redirect, url_for, jsonify, make_response
from openpyxl import load_workbook
import sqlite3
import pandas as pd
import uuid
from minio import Minio
from dotenv import load_dotenv
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from sqlalchemy import update, desc
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from jose import JWTError, jwt
from functools import wraps
import json
import numpy as np
import math
import random
 
load_dotenv()

app = Flask(__name__)
app.config[
    "SQLALCHEMY_DATABASE_URI"
] = "postgresql://sonakshi:sonakshi@localhost:5432/hiringapp"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = True
app.config['SECRET_KEY']='something'
db = SQLAlchemy(app)
migrate = Migrate(app, db)
from models import *
from sqlalchemy import create_engine, select, MetaData, Table, and_

engine = create_engine("postgresql://sonakshi:sonakshi@localhost:5432/hiringapp")
metadata = MetaData(bind=None)
# minioClient = Minio(
#         "play.min.io",
#         access_key="Q3AM3UQ867SPQQA43P2F",
#         secret_key="zuf+tfteSlswRu7BJ86wekitnifILbZam1KYY3TG",
#         secure=True,
#     )
# found = minioClient.bucket_exists("resume")
# if not found:
#     minioClient.make_bucket("resume")
# else:
#     print("Bucket 'resume' already exists")

#------------------Auth and roles access implementation--------------
def token_required(f):
	@wraps(f)
	def decorated(*args, **kwargs):
		token = request.cookies.get('token') #for python and js compatibilty
		if token:
			try:
				#print(token.split()[1])
				data = jwt.decode(token, app.config['SECRET_KEY'], 'HS256')
				current_user = Users.query\
					.filter_by(id = data['id'])\
					.first()
			except:
				return jsonify({
					'message' : 'Token is invalid !!'
				}), 401
		if not token:
			return jsonify({'message' : 'Token is missing !!'}), 401


		return f(current_user, *args, **kwargs)

	return decorated

def token_forwarder(f):
	@wraps(f)
	def decorated(*args, **kwargs):
		token = request.cookies.get('token') #for python and js compatibilty
		if token:#'Authorization' in request.headers:
			#token = #request.headers['Authorization']
			try:
				data = jwt.decode(token, app.config['SECRET_KEY'], 'HS256')
				current_user = Users.query\
    				.filter_by(id = data['id'])\
    				.first()
			except: return render_template('login.html', user=None)
				# return jsonify({
    			# 	'message' : 'Token is invalid !!'
    			# }), 401
		if not token:
			current_user=None


		return f(current_user, *args, **kwargs)

	return decorated

def pmoOnly(f):
	@wraps(f)
	def decorated(*args, **kwargs):
		token = request.cookies.get('token') #for python and js compatibilty
		if token:
			try:
				#print(token.split()[1])
				data = jwt.decode(token, app.config['SECRET_KEY'], 'HS256')
				current_user = Users.query\
					.filter_by(id = data['id'])\
					.first()
				if(current_user.role != 'pmo'):
					return jsonify({
						'message' : 'Unauthorised access to role PMO !!'
					}), 401
					current_user=None
			except:
				return render_template('login.html', user=None)
		if not token:
			return jsonify({'message' : 'Token is missing !!'}), 401

		return f(current_user, *args, **kwargs)

	return decorated

def tacOnly(f):
	@wraps(f)
	def decorated(*args, **kwargs):
		token = request.cookies.get('token') #for python and js compatibilty
		if token:
			try:
				#print(token.split()[1])
				data = jwt.decode(token, app.config['SECRET_KEY'], 'HS256')
				current_user = Users.query\
					.filter_by(id = data['id'])\
					.first()
				if(current_user.role != 'tac'):
					return jsonify({
						'message' : 'Unauthorised access to role TAC team !!'
					}), 401
					current_user=None
			except:
				return render_template('login.html', user=None)
		if not token:
			return jsonify({'message' : 'Token is missing !!'}), 401


		return f(current_user, *args, **kwargs)

	return decorated

def interviewerOnly(f):
	@wraps(f)
	def decorated(*args, **kwargs):
		token = request.cookies.get('token') #for python and js compatibilty
		if token:
			try:
				#print(token.split()[1])
				data = jwt.decode(token, app.config['SECRET_KEY'], 'HS256')
				current_user = Users.query\
					.filter_by(id = data['id'])\
					.first()
				if(current_user.role != 'interviewer'):
					return jsonify({
						'message' : 'Unauthorised access to role Interviewer !!'
					}), 401
					current_user=None
			except:
				return render_template('login.html', user=None)
		if not token:
			return jsonify({'message' : 'Token is missing !!'}), 401


		return f(current_user, *args, **kwargs)

	return decorated

def hrOnly(f):
	@wraps(f)
	def decorated(*args, **kwargs):
		token = request.cookies.get('token') #for python and js compatibilty
		if token:
			try:
				#print(token.split()[1])
				data = jwt.decode(token, app.config['SECRET_KEY'], 'HS256')
				current_user = Users.query\
					.filter_by(id = data['id'])\
					.first()
				if(current_user.role != 'hr'):
					return jsonify({
						'message' : 'Unauthorised access to role HR team !!'
					}), 401
					current_user=None
			except:
				return render_template('login.html', user=None)
		if not token:
			return jsonify({'message' : 'Token is missing !!'}), 401


		return f(current_user, *args, **kwargs)

	return decorated

def Admin(f):
	@wraps(f)
	def decorated(*args, **kwargs):
		token = request.cookies.get('token') #for python and js compatibilty
		if token:
			try:
				#print(token.split()[1])
				data = jwt.decode(token, app.config['SECRET_KEY'], 'HS256')
				current_user = Users.query\
					.filter_by(id = data['id'])\
					.first()
				if(current_user.role != 'admin'):
					return jsonify({
						'message' : 'Unauthorised access to role Admin !!'
					}), 401
					current_user=None
			except:
				return render_template('login.html', user=None)
		if not token:
			return jsonify({'message' : 'Token is missing !!'}), 401


		return f(current_user, *args, **kwargs)

	return decorated

#example of pmo role specific routing
@app.route('/test', methods=['GET'])
@pmoOnly
def getPMO(current_user):
	if(current_user):
		return jsonify({'users': current_user.name})
	else:
		return jsonify({'users': None})

@app.route("/logout", methods=["GET"])
def removeToken():
	resp = make_response(redirect(url_for("main")))
	resp.delete_cookie('token')
	return resp
# an example of logged in admin route to access all data
@app.route('/users', methods =['GET'])
@token_required
def get_all_users(current_user):
	users = Users.query.all()
	output = []
	for user in users:
		output.append({
			'id': user.id,
			'name' : user.name,
			'email' : user.email,
			'role': user.role
		})

	return jsonify({'users': output})

@app.route('/loginToken', methods =['POST'])
def loginToken():
	auth = request.form
	print(auth.get('email'), auth.get('password'))
	if not auth or not auth.get('email') or not auth.get('password'):
		return make_response(
			'Could not verify',
			401,
			{'WWW-Authenticate' : 'Basic realm ="Login required !!"'}
		)

	user = Users.query\
		.filter_by(email = auth.get('email'))\
		.first()

	if not user:
		return make_response(
			'Could not verify',
			401,
			{'WWW-Authenticate' : 'Basic realm ="User does not exist !!"'}
		)

	if check_password_hash(user.password, auth.get('password')):
		token = jwt.encode({
			'id': user.id,
			'exp' : datetime.utcnow() + timedelta(minutes = 30)
		}, app.config['SECRET_KEY'], 'HS256')
	resp = make_response(redirect(url_for("main")))#make_response(jsonify({'token' : token, 'user':user.role}), 201)
	resp.set_cookie('token', token)
	return resp

	return make_response(
		'Could not verify',
		403,
		{'WWW-Authenticate' : 'Basic realm ="Wrong Password !!"'}
	)

@app.route('/createUser', methods =['POST'])
@Admin
def createUser(user):
	data = request.form
	name, email, role = data.get('name'), data.get('email'), data.get('role')
	password = data.get('password')
	user = Users.query\
		.filter_by(email = email)\
		.first()
	if not user:
		user = Users(
			id = str(uuid.uuid4()),
			name = name,
			email = email,
			password = generate_password_hash(password),
			role = role
		)
		db.session.add(user)
		db.session.commit()

		return make_response('Successfully registered.', 201)
	else:
		return make_response('User already exists', 202)


#--------------------------------------------------------------------


def get_abstract_data():
    result = roles.query.order_by(desc(roles.modified_at)).all()
    nums = {}
    col={
        "green":1,
        "yellow":2,
        "red":3,
        "gray":4
    }
    for i in result:
        nums[i.tsin_id.strip()] = {
            "stage1": 0,
            "stage2": 0,
            "stage3": 0,
            "stage4": 0,
            "stage5": 0,
            "stage6": 0,
            "stage7": 0,
            "stage8": 0,
            "stage9": 0,
        }
    result2 = candidates.query.all()
    list_of_phone = []
    for i in result2:
        list_of_phone.append(i.phone)
    d = {}
    final_color={}
    list_of_tsin=[]
    for i in result:
        list_of_tsin.append(i.tsin_id.strip())
    for i in result2:
        if i.tsin_id.strip() in list_of_tsin:
            if i.status=="inactive":
                color="gray"
            else:
                if i.request_raised_date:
                    difference=datetime.now()-i.request_raised_date
                elif i.resume_screened_date:
                    difference=datetime.now()-i.resume_screened_date
                else:
                    difference=timedelta(days=0)# logical error heree if both cases are false difference is not made at all
                if i.current_stage.strip() == "Resume Screened for Interview":
                    nums[i.tsin_id.strip()]["stage1"] += 1
                    if difference.days<=8:
                        color="green"
                    elif difference.days>8 and difference.days<10:
                        color="yellow"
                    else:
                        color="red"
                elif i.current_stage.strip() == "L1 Interview Complete":
                    nums[i.tsin_id.strip()]["stage2"] += 1
                    if difference.days<=13:
                        color="green"
                    elif difference.days>13 and difference.days<15:
                        color="yellow"
                    else:
                        color="red"
                elif i.current_stage.strip() == "L2 Interview Complete":
                    nums[i.tsin_id.strip()]["stage3"] += 1
                    if difference.days<=16:
                        color="green"
                    elif difference.days>16 and difference.days<18:
                        color="yellow"
                    else:
                        color="red"
                elif i.current_stage.strip() == "L3 Interview Complete":
                    nums[i.tsin_id.strip()]["stage4"] += 1
                    if difference.days<=21:
                        color="green"
                    elif difference.days>21 and difference.days<23:
                        color="yellow"
                    else:
                        color="red"
                elif i.current_stage.strip() == "Offer RollOut":
                    nums[i.tsin_id.strip()]["stage5"] += 1
                    if difference.days<=23:
                        color="green"
                    elif difference.days>23 and difference.days<25:
                        color="yellow"
                    else:
                        color="red"
                elif i.current_stage.strip() == "Buddy Assignment":
                    nums[i.tsin_id.strip()]["stage6"] += 1
                    if difference.days<=25:
                        color="green"
                    elif difference.days>25 and difference.days<27:
                        color="yellow"
                    else:
                        color="red"
                elif (
                    i.current_stage.strip() == "Candidate Joined"
                    or i.current_stage.strip() == "Candidate Dropout"
                ):
                    nums[i.tsin_id.strip()]["stage7"] += 1
                    color="gray"
                elif i.current_stage.strip() == "Resume Rejected":
                    nums[i.tsin_id.strip()]["stage8"] += 1
                    color="gray"
                elif i.current_stage.strip() == "Profile Upload":
                    nums[i.tsin_id.strip()]["stage9"] += 1
                    if difference.days<=5:
                        color="green"
                    elif difference.days>5 and difference.days<5:
                        color="yellow"
                    else:
                        color="red"
            if i.tsin_id.strip() in d:
                d[i.tsin_id.strip()].append(
                    {
                        "id": i.id,
                        "candidate_name": i.candidate_name,
                        "pan": i.pan,
                        "candidate_email": i.candidate_email,
                        "current_stage": i.current_stage,
                        "request_raised_date": i.request_raised_date,
                        "tsin_opened_date": i.tsin_opened_date,
                        "resume_screened_date": i.resume_screened_date,
                        "l1_interview_date": i.l1_interview_date,
                        "l1_interviewer": i.l1_interviewer,
                        "l2_interview_date": i.l2_interview_date,
                        "l2_interviewer": i.l2_interviewer,
                        "l3_interview_date": i.l3_interview_date,
                        "l3_interviewer": i.l3_interviewer,
                        "offer_rollout_date": i.offer_rollout_date,
                        "joining_date": i.joining_date,
                        "buddy_assignment_date": i.buddy_assignment_date,
                        "buddy_name": i.buddy_name,
                        "candidate_dropout_date": i.candidate_dropout_date,
                        "candidate_dropout_reason": i.candidate_dropout_reason,
                        "resume": i.resume,
                        "phone": i.phone,
                        "current_location": i.current_location,
                        "current_company": i.current_company,
                        "experience": i.experience,
                        "status": i.status,
                        "color":color
                    }
                )
            else:
                d[i.tsin_id.strip()] = [
                    {
                        "id": i.id,
                        "candidate_name": i.candidate_name,
                        "pan": i.pan,
                        "candidate_email": i.candidate_email,
                        "current_stage": i.current_stage,
                        "request_raised_date": i.request_raised_date,
                        "tsin_opened_date": i.tsin_opened_date,
                        "resume_screened_date": i.resume_screened_date,
                        "l1_interview_date": i.l1_interview_date,
                        "l1_interviewer": i.l1_interviewer,
                        "l2_interview_date": i.l2_interview_date,
                        "l2_interviewer": i.l2_interviewer,
                        "l3_interview_date": i.l3_interview_date,
                        "l3_interviewer": i.l3_interviewer,
                        "offer_rollout_date": i.offer_rollout_date,
                        "joining_date": i.joining_date,
                        "buddy_assignment_date": i.buddy_assignment_date,
                        "buddy_name": i.buddy_name,
                        "candidate_dropout_date": i.candidate_dropout_date,
                        "candidate_dropout_reason": i.candidate_dropout_reason,
                        "resume": i.resume,
                        "phone": i.phone,
                        "current_location": i.current_location,
                        "current_company": i.current_company,
                        "experience": i.experience,
                        "status":i.status,
                        "color":color
                    }
                ]
        if i.tsin_id in final_color.keys():
            if col[final_color[i.tsin_id]]>col[color]:
                final_color[i.tsin_id]=color
        else:
            final_color[i.tsin_id]=color

    return result, d, nums, list_of_phone, final_color


def total_offers():
    result = candidates.query.with_entities(
        candidates.tsin_id,
        candidates.offer_rollout_date,
        candidates.candidate_dropout_date,
    ).all()

    final = {}
    dropout_final = {}
    for i in result:
        j = roles.query.with_entities(roles.role).filter_by(tsin_id=i.tsin_id).all()
        print(j)
        if i.offer_rollout_date != None:
            time = i.offer_rollout_date
            now = datetime.now()
            difference = now - time
            if j[0][0] in final.keys():
                print(".....", final[j[0][0]])
                final[j[0][0]].append(difference.days)
            else:
                final[j[0][0]] = [difference.days]
        if i.candidate_dropout_date != None:
            time = i.candidate_dropout_date
            now = datetime.now()
            difference = now - time
            if j[0][0] in dropout_final.keys():
                dropout_final[j[0][0]].append(difference.days)
            else:
                dropout_final[j[0][0]] = [difference.days]
    print(final)

    return final, dropout_final


def get_new_profiles():
    result = candidates.query.with_entities(
        candidates.tsin_id, candidates.created_at
    ).all()
    final = {}
    for i in result:
        j = roles.query.with_entities(roles.role).filter_by(tsin_id=i.tsin_id).all()
        print(j)
        now = datetime.now()
        difference = now - i.created_at
        print(j[0][0])
        if j[0][0] in final.keys():
            final[j[0][0]].append(difference.days)
        else:
            final[j[0][0]] = [difference.days]
    return final


def get_role_wise_profiles():
    result = candidates.query.with_entities(
        candidates.tsin_id, candidates.current_stage
    ).all()
    final = {}
    roless=roles.query.distinct(roles.role).all()
    print(roless)
    final_roles=[]
    profile_upload, resume_screened,l1_interview, l2_interview, l3_interview, offer_rollout, buddy_assignment, candidate_joined=[0]*len(roless), [0]*len(roless), [0]*len(roless) , [0]*len(roless), [0]*len(roless), [0]*len(roless), [0]*len(roless), [0]*len(roless)
    for i in roless:
        final_roles.append(i.role)
    for i in result:
        j = roles.query.with_entities(roles.role).filter_by(tsin_id=i.tsin_id).all()
        print(j)
        z=final_roles.index(j[0][0])
        print(j,z)
        if i.current_stage == "Profile Upload":
            profile_upload[z]+=1
        elif i.current_stage == "Resume Screened for Interview":
            resume_screened[z]+=1
        elif i.current_stage == "L1 Interview Complete":
            l1_interview[z]+=1
        elif i.current_stage == "L2 Interview Complete":
            l2_interview[z]+=1
        elif i.current_stage == "L3 Interview Complete":
            l3_interview[z]+=1
        elif i.current_stage == "Offer RollOut":
            offer_rollout[z]+=1
        elif i.current_stage == "Buddy Assignment":
            buddy_assignment[z]+=1
        elif (
            i.current_stage == "Candidate Joined"
            or i.current_stage == "Candidate Dropout"
        ):
            candidate_joined[z]+=1
    final2 = []
    for i in final:
        final3 = [i] + final[i]
        final2.append(final3)
    print(final2)

    return final_roles, profile_upload,resume_screened,l1_interview,l2_interview,l3_interview,offer_rollout,buddy_assignment,candidate_joined


def get_detailed_data():
    result = roles.query.all()
    result2 = candidates.query.all()
    roless = {}
    for i in result:
        roless[i.tsin_id.strip()] = {
            "role": i.role,
            "chapter": i.chapter,
            "squad": i.squad,
            "demand_type": i.demand_type,
            "tribe": i.tribe,
            "snow_id": i.snow_id,
        }
    d = []
    for i in result2:
        d.append(
            {
                "tsinid": i.tsin_id,
                "chapter": roless[i.tsin_id.strip()]["chapter"],
                "squad": roless[i.tsin_id.strip()]["squad"],
                "demand_type": roless[i.tsin_id.strip()]["demand_type"],
                "tribe": roless[i.tsin_id.strip()]["tribe"],
                "snow_id": roless[i.tsin_id.strip()]["snow_id"],
                "role": roless[i.tsin_id.strip()]["role"],
                "id": i.id,
                "candidate_name": i.candidate_name,
                "pan": i.pan,
                "candidate_email": i.candidate_email,
                "current_stage": i.current_stage,
                "request_raised_date": i.request_raised_date,
                "tsin_opened_date": i.tsin_opened_date,
                "resume_screened_date": i.resume_screened_date,
                "l1_interview_date": i.l1_interview_date,
                "l1_interviewer": i.l1_interviewer,
                "l2_interview_date": i.l2_interview_date,
                "l2_interviewer": i.l2_interviewer,
                "l3_interview_date": i.l3_interview_date,
                "l3_interviewer": i.l3_interviewer,
                "offer_rollout_date": i.offer_rollout_date,
                "joining_date": i.joining_date,
                "buddy_assignment_date": i.buddy_assignment_date,
                "buddy_name": i.buddy_name,
                "candidate_dropout_date": i.candidate_dropout_date,
                "candidate_dropout_reason": i.candidate_dropout_reason,
                "resume": i.resume,
                "phone": i.phone,
                "current_location": i.current_location,
                "current_company": i.current_company,
                "experience": i.experience,
            }
        )
    return d



@app.route("/")
@token_forwarder
def main(user):
    if(user):
        rolesss=user.role
    else:
        rolesss=None
    result, d, nums, phones, colors= get_abstract_data()
    roless=role.query.all()
    chapters=chapter.query.all()
    squads=squad.query.all()
    tribes=tribe.query.all()
    return render_template(
        "candidates.html", roles=result, candidates=d, nums=nums, phones=phones, final_colors=colors, roless=roless, squads=squads, tribes=tribes, chapters=chapters, user=rolesss
    )


@app.route("/upload-dataset", methods=["GET", "POST"])
def uploadDs():
    f = request.files["file"]  # File input
    if not f:
        return "No file attached"

    global filename
    filename = f.filename  # changing global value of filename

    path = "{}/{}".format("static", filename)
    f.save(path)
    x = path.split(".")[-1]

    # reading filedata start
    if x == "xlsx":
        new_wb = load_workbook(path)
        Dataframe = pd.read_excel(new_wb, engine="openpyxl")
    elif x == "csv":
        Dataframe = pd.read_csv(path, encoding="ISO-8859-1")
    else:
        return "Please upload the file in xlsx or csv only"
    # reading filedata end

    dict_of_records = Dataframe.to_dict(orient="record")

    for i in dict_of_records:
        current_stage="Empty"
        candidate_dropout_date = str(i["Candidate Dropout"])
        candidate_joined_date = str(i["Candidate Joined"])
        if candidate_dropout_date == "nan" or candidate_dropout_date == "NaT":
            candidate_dropout_date = None
        else:
            current_stage = "Candidate Dropout"

        if candidate_joined_date == "nan" or candidate_joined_date == "NaT":
            candidate_joined_date = None
        else:
            if current_stage=="Empty":
                current_stage = "Candidate Joined"

        buddy_assignment_date = str(i["Buddy Assignment "])
        if buddy_assignment_date == "nan" or buddy_assignment_date == "NaT":
            buddy_assignment_date = None
        else:
            if current_stage=="Empty":
                current_stage = "Buddy Assignment"

        joining_date = str(i["Joining Date"])
        if joining_date == "nan" or joining_date == "NaT":
            joining_date = None
        else:
            if current_stage=="Empty":
                current_stage = "Offer RollOut"

        offer_rollout_date = str(i["Offer RollOut "])
        if offer_rollout_date == "nan" or offer_rollout_date == "NaT":
            offer_rollout_date = None
        else:
            if current_stage=="Empty":
                current_stage = "Offer RollOut"

        l3_interview_date = str(i["L3 Interview Complete"])
        if l3_interview_date == "nan" or l3_interview_date == "NaT":

            l3_interview_date = None
        else:
            if current_stage=="Empty":
                current_stage = "L3 Interview Complete"

        l2_interview_date = str(i["L2 Interview Complete"])
        if l2_interview_date == "nan" or l2_interview_date == "NaT":
            l2_interview_date = None
        else:
            if current_stage=="Empty":
                current_stage = "L2 Interview Complete"

        l1_interview_date = str(i["L1 Interview Complete"])
        if l1_interview_date == "nan" or l1_interview_date == "NaT":
            current_stage = "Resume Screened for Interview"
            l1_interview_date = None
        else:
            if current_stage=="Empty":
                current_stage = "L1 Interview Complete"

        resume_screened_date = str(i["Resume Screened"])
        if resume_screened_date == "nan" or resume_screened_date == "NaT":
            resume_screened_date = None
        else:
            if current_stage=="Empty":
                current_stage = "Profile Upload"

        tsin_opened_date = str(i["TSIN Opened"])
        if tsin_opened_date == "nan" or tsin_opened_date == "NaT":
            tsin_opened_date = None
        else:
            if current_stage=="Empty":
                current_stage = "Profile Upload"

        request_raised_date = str(i["Resume Shared"])
        if request_raised_date == "nan" or request_raised_date == "NaT":
            request_raised_date = None
        else:
            if current_stage=="Empty":
                current_stage = "Profile Upload"

        existing_roles = roles.query.filter_by(tsin_id=i["TSIN ID"].strip()).all()
        existing_emails = candidates.query.with_entities(
            candidates.candidate_email, candidates.tsin_id
        ).all()
        print(existing_emails)
        if len(existing_roles)==0:
            db.session.add(
                roles(
                    tsin_id=str(i["TSIN ID"].strip()),
                    role=str(i["Role "]),
                    chapter=str(i["Chapter"]),
                    squad=str(i["Squad"]),
                    demand_type=str(i["Type of Demand"]),
                    tribe=str(i["Tribe"]),
                    snow_id=str(i["Snow ID"]),
                    aspd_date=str(i["APSD Date"]),
                    status="active",
                    created_at=datetime.now(),
                    modified_at=datetime.now(),
                )
            )
            db.session.commit()
        else:
            existing_roles[0].tsin_id=str(i["TSIN ID"].strip())
            existing_roles[0].role=str(i["Role "])
            existing_roles[0].chapter=str(i["Chapter"])
            existing_roles[0].squad=str(i["Squad"])
            existing_roles[0].demand_type=str(i["Type of Demand"])
            existing_roles[0].tribe=str(i["Tribe"])
            existing_roles[0].snow_id=str(i["Snow ID"])
            existing_roles[0].aspd_date=str(i["APSD Date"])
            existing_roles[0].status="active"
            db.session.commit()
        z = [str(i["Cadidate Email"]), str(i["TSIN ID"].strip())]
        if tuple(z) in existing_emails:
            print("yes")
            z = candidates.query.filter_by(
                candidate_email=str(i["Cadidate Email"]),
                tsin_id=str(i["TSIN ID"].strip()),
            ).update(
                {
                    "tsin_id": str(i["TSIN ID"].strip()),
                    "candidate_name": str(i["Candidate Name "]),
                    "pan": str(i["PAN Number"]),
                    "candidate_email": str(i["Cadidate Email"]),
                    "current_stage": current_stage,
                    "request_raised_date": request_raised_date,
                    "tsin_opened_date": tsin_opened_date,
                    "resume_screened_date": resume_screened_date,
                    "l1_interview_date": l1_interview_date,
                    "l1_interviewer": str(i["L1 Interviewer"]),
                    "l2_interview_date": l2_interview_date,
                    "l2_interviewer": str(i["L2 interviewer"]),
                    "l3_interview_date": l3_interview_date,
                    "l3_interviewer": str(i["L3 Interviewer"]),
                    "offer_rollout_date": offer_rollout_date,
                    "joining_date": joining_date,
                    "buddy_assignment_date": buddy_assignment_date,
                    "buddy_name": str(i["Buddy Name"]),
                    "candidate_dropout_date": candidate_dropout_date,
                    "candidate_dropout_reason": str(i["Dropout Reason"]),
                    "modified_at": datetime.now(),
                    "status": "active",
                }
            )
            # z.tsin_id = str(i["TSIN ID"].strip())
            # z.candidate_name = str(i["Candidate Name "])
            # z.pan = str(i["PAN Number"])
            # z.candidate_email = str(i["Cadidate Email"])
            # z.current_stage = current_stage
            # z.request_raised_date = request_raised_date
            # z.status="active"
            db.session.commit()

        else:
            db.session.add(
                candidates(
                    tsin_id=str(i["TSIN ID"].strip()),
                    candidate_name=str(i["Candidate Name "]),
                    pan=str(i["PAN Number"]),
                    candidate_email=str(i["Cadidate Email"]),
                    current_stage=current_stage,
                    request_raised_date=request_raised_date,
                    tsin_opened_date=tsin_opened_date,
                    resume_screened_date=resume_screened_date,
                    l1_interview_date=l1_interview_date,
                    l1_interviewer=str(i["L1 Interviewer"]),
                    l2_interview_date=l2_interview_date,
                    l2_interviewer=str(i["L2 interviewer"]),
                    l3_interview_date=l3_interview_date,
                    l3_interviewer=str(i["L3 Interviewer"]),
                    offer_rollout_date=offer_rollout_date,
                    joining_date=joining_date,
                    buddy_assignment_date=buddy_assignment_date,
                    buddy_name=str(i["Buddy Name"]),
                    candidate_dropout_date=candidate_dropout_date,
                    candidate_dropout_reason=str(i["Dropout Reason"]),
                    resume="",
                    resume_screened_remarks="",
                    l1_interview_result="",
                    l1_interview_remarks="",
                    l2_interview_remarks="",
                    l2_interview_result="",
                    l3_interview_result="",
                    l3_interview_remarks="",
                    phone="",
                    current_location="",
                    current_company="",
                    experience="",
                    candidate_joined_date=candidate_joined_date,
                    created_at=datetime.now(),
                    modified_at=datetime.now(),
                    resume_remarks="",
                    status="active",
                    l1_completion=None,
                    l2_completion=None,
                    l3_completion=None
                )
            )
            db.session.commit()
    print("added to db")
    return redirect(url_for("main"))


@app.route("/upload-profile", methods=["GET", "POST"])
def profileUpload():
    cand_id = request.form["candidate_id"]
    phone = request.form["phone"]
    company = request.form["company"]
    experience = request.form["experience"]
    location = request.form["location"]
    resume = request.files["resume"]
    result2 = (
        candidates.query.with_entities(candidates.candidate_name)
        .filter_by(id=cand_id)
        .all()
    )
    if resume:
        resume.save("static/resume")
        # minioClient.fput_object(
        #     "resume",
        #     "resume" + str(result2[0][0]).strip(),
        #     "static/resume",
        # )
        print(
            "resume is successfully uploaded as "
            "object 'resume" + str(result2[0][0]).strip() + "' to bucket 'resume'."
        )

    db.session.query(candidates).filter(candidates.id == cand_id).update(
        {
            "phone": phone,
            "current_location": location,
            "current_company": company,
            "experience": experience,
            "modified_at": datetime.now(),
        }
    )
    db.session.commit()
    return redirect(url_for("main"))


# @app.route("/download-resume", methods=["GET", "POST"])
# def downloadresume():
#     id = request.form.get("id")
#     res = (
#         candidates.query.with_entities(candidates.candidate_name).filter_by(id=id).all()
#     )
#     z = minioClient.get_object("resume", "resume" + res[0][0])
#     minioClient.fget_object("resume", "resume" + res[0][0], "static/resume.pdf")
#     return z.data


@app.route("/upload-candidates", methods=["GET", "POST"])
@token_forwarder
def candidates1(user):
    #print(token)
	if(user):
		role=user.role
	else:
		role=None
	return render_template("index.html",user=role)


@app.route("/update-stage", methods=["GET", "POST"])
def updateStage():
    cand_id = request.form.get("candidate_id")
    stage = request.form.get("stage")
    resume_remarks = request.form.get("resume_remarks")
    update_time = request.form.get("stageupdatetime")
    result = request.form.get("result")
    l1_remarks = request.form.get("l1_remarks")
    l2_interviewer_name = request.form.get("l2_interviewer_name")
    l2_interview_date = request.form.get("l2_interview_date")
    l2_remarks = request.form.get("l2_remarks")
    l3_interviewer_name = request.form.get("l3_interviewer_name")
    l3_interview_date = request.form.get("l3_interview_date")
    l3_remarks = request.form.get("l3_remarks")
    joining_date = request.form.get("joining_date")
    buddy_name = request.form.get("buddy_name")
    joining_stage_time = request.form.get("joiningstageupdatetime")
    dropout_stage_time = request.form.get("dropoutstageupdatetime")
    dropout_reason = request.form.get("dropout_reason")
    print(update_time, result, resume_remarks)
    candidate=candidates.query.filter_by(id = cand_id).first()
    if stage == "resumescreened":
        if result == "passed":
            print('abc')
            db.session.query(candidates).filter(candidates.id == cand_id).update(
                {
                    "current_stage": "Resume Screened for Interview",
                    "resume_screened_date": update_time,
                    "resume_remarks": resume_remarks,
                    "modified_at": datetime.now(),
                }
            )
        else:
            db.session.query(candidates).filter(candidates.id == cand_id).update(
                {
                    "current_stage": "Resume Rejected",
                    "resume_screened_date": update_time,
                    "modified_at": datetime.now(),
                }
            )
        db.session.commit()
    if stage == "l1complete":
        if result == "passed":
            if candidate.l1_interview_date!=None:
                db.session.query(candidates).filter(candidates.id == cand_id).update(
                    {
                        "current_stage": "L1 Interview Complete",
                        "l1_completion": update_time,
                        "l1_interview_result": result,
                        "l1_interview_remarks": l1_remarks,
                        "l2_interviewer": l2_interviewer_name,
                        "l2_interview_date": l2_interview_date,
                        "modified_at": datetime.now(),
                    }
                )
            else:
                db.session.query(candidates).filter(candidates.id == cand_id).update(
                    {
                        "current_stage": "L1 Interview Complete",
                        "l1_interview_date":update_time,
                        "l1_completion": update_time,
                        "l1_interview_result": result,
                        "l1_interview_remarks": l1_remarks,
                        "l2_interviewer": l2_interviewer_name,
                        "l2_interview_date": l2_interview_date,
                        "modified_at": datetime.now(),
                    }
                )
        else:
            db.session.query(candidates).filter(candidates.id == cand_id).update(
                {
                    "current_stage": "Resume Rejected",
                    "l1_completion": update_time,
                    "l1_interview_result": result,
                    "l1_interview_remarks": l1_remarks,
                    "modified_at": datetime.now(),
                }
            )
        db.session.commit()
    elif stage == "l2complete":
        if result == "passed":
            if candidate.l2_interview_date!=None:
                db.session.query(candidates).filter(candidates.id == cand_id).update(
                    {
                        "current_stage": "L2 Interview Complete",
                        "l2_completion": update_time,
                        "l2_interview_result": result,
                        "l2_interview_remarks": l2_remarks,
                        "l3_interviewer": l3_interviewer_name,
                        "l3_interview_date": l3_interview_date,
                    }
                )
            else:
                db.session.query(candidates).filter(candidates.id == cand_id).update(
                    {
                        "current_stage": "L2 Interview Complete",
                        "l2_completion": update_time,
                        "l2_interview_date":update_time,
                        "l2_interview_result": result,
                        "l2_interview_remarks": l2_remarks,
                        "l3_interviewer": l3_interviewer_name,
                        "l3_interview_date": l3_interview_date,
                    }
                )
        else:
            db.session.query(candidates).filter(candidates.id == cand_id).update(
                {
                    "current_stage": "Resume Rejected",
                    "l2_completion": update_time,
                    "l2_interview_result": result,
                    "l2_interview_remarks": l2_remarks,
                    "modified_at": datetime.now(),
                }
            )
        db.session.commit()
    elif stage == "l3complete":
        if result == "passed":
            if candidate.l3_interview_date!=None:
                db.session.query(candidates).filter(candidates.id == cand_id).update(
                    {
                        "current_stage": "L3 Interview Complete",
                        "l3_completion": update_time,
                        "l3_interview_result": result,
                        "l3_interview_remarks": l3_remarks,
                        "modified_at": datetime.now(),
                    }
                )
            else:
                db.session.query(candidates).filter(candidates.id == cand_id).update(
                    {
                        "current_stage": "L3 Interview Complete",
                        "l3_interview_date":update_time,
                        "l3_completion": update_time,
                        "l3_interview_result": result,
                        "l3_interview_remarks": l3_remarks,
                        "modified_at": datetime.now(),
                    }
                )
        else:
            db.session.query(candidates).filter(candidates.id == cand_id).update(
                {
                    "current_stage": "Resume Rejected",
                    "l3_completion": update_time,
                    "l3_interview_result": result,
                    "l3_interview_remarks": l3_remarks,
                    "modified_at": datetime.now(),
                }
            )
        db.session.commit()
    elif stage == "offer":
        db.session.query(candidates).filter(candidates.id == cand_id).update(
            {
                "current_stage": "Offer RollOut",
                "offer_rollout_date": update_time,
                "joining_date": joining_date,
                "modified_at": datetime.now(),
            }
        )
        db.session.commit()
    elif stage == "buddy":
        db.session.query(candidates).filter(candidates.id == cand_id).update(
            {
                "current_stage": "Buddy Assignment",
                "buddy_assignment_date": update_time,
                "buddy_name": buddy_name,
                "modified_at": datetime.now(),
            }
        )
        db.session.commit()
    elif stage == "finalcandidate":
        if joining_stage_time and dropout_stage_time:
            return "Please choose appropriate values - Joining and dropout cannot come together!"
        elif joining_stage_time:
            db.session.query(candidates).filter(candidates.id == cand_id).update(
                {
                    "current_stage": "Candidate Joined",
                    "candidate_joined_date": joining_stage_time,
                    "modified_at": datetime.now(),
                }
            )
        elif dropout_stage_time:
            db.session.query(candidates).filter(candidates.id == cand_id).update(
                {
                    "current_stage": "Candidate Dropout",
                    "candidate_dropout_date": dropout_stage_time,
                    "candidate_dropout_reason": dropout_reason,
                    "modified_at": datetime.now(),
                }
            )
        db.session.commit()
    return redirect(url_for("main"))


@app.route("/add-tsin", methods=["GET", "POST"])
@token_forwarder
def tsinform(user):
    #print(token)
    result = roles.query.with_entities(roles.tsin_id).all()
    roless=role.query.all()
    chapters=chapter.query.all()
    squads=squad.query.all()
    tribes=tribe.query.all()
    final = []
    for i in result:
        final.append(i[0])
    if(user):
        Userrole=user.role
    else:
        Userrole=None
    return render_template("tsin_form.html", final=final, roles=roless, chapters=chapters, squads=squads, tribes=tribes, user=Userrole)


@app.route("/detailed-view", methods=["GET", "POST"])
@token_forwarder
def detailed(user):
    #print(token)
    tsin = request.form.get("val1")
    role = request.form.get("val2")
    chapter = request.form.get("val3")
    squad = request.form.get("val4")
    cname = request.form.get("val5")
    stage = request.form.get("val6")
    print(tsin, role, chapter, squad, cname)
    z = get_detailed_data()
    final = []
    for i in z:
        if tsin:
            if i["tsinid"].lower().strip() == tsin.lower().strip():
                if i not in final:
                    final.append(i)
        if role:
            if i["role"].lower().strip() == role.lower().strip():
                if i not in final:
                    final.append(i)
        if chapter:
            if i["chapter"].lower().strip() == chapter.lower().strip():
                if i not in final:
                    final.append(i)
        if squad:
            if i["squad"].lower().strip() == squad.lower().strip():
                if i not in final:
                    final.append(i)
        if cname:
            if i["candidate_name"].lower().strip() == cname.lower().strip():
                if i not in final:
                    final.append(i)
        if stage:
            if i["current_stage"].lower().strip() == stage.lower().strip():
                if i not in final:
                    final.append(i)
    if(user):
        role=user.role
    else:
        role=None
    if final==[]:
        print('abc')
        return render_template("detailed.html", final=z, user=role)
    else:
        print('def')
        return render_template("filtered_details.html", final=final)

@app.route("/activate-role/<tsin>", methods=["GET", "POST"])
def activate_role(tsin):
    db.session.query(roles).filter(roles.tsin_id == tsin).update(
        {
            "status": "active",
            "modified_at": datetime.now(),
        }
    )
    db.session.commit()
    return redirect(url_for("main"))

@app.route("/delete-role/<tsin>", methods=["GET", "POST"])
def delete_role(tsin):
    db.session.query(roles).filter(roles.tsin_id == tsin).update(
        {
            "status": "inactive",
            "modified_at": datetime.now(),
        }
    )
    db.session.commit()
    return redirect(url_for("main"))

@app.route("/activate-profile/<id>", methods=["GET", "POST"])
def activate_profile(id):
    db.session.query(candidates).filter(candidates.id == id).update(
        {
            "status": "active",
            "modified_at": datetime.now(),
        }
    )
    db.session.commit()
    return redirect(url_for("main"))

@app.route("/delete-profile/<id>", methods=["GET", "POST"])
def delete_profile(id):
    db.session.query(candidates).filter(candidates.id == id).update(
        {
            "status": "inactive",
            "modified_at": datetime.now(),
        }
    )
    db.session.commit()
    return redirect(url_for("main"))


@app.route("/new-position", methods=["GET", "POST"])
def newposition():
    total=request.form.get("total")
    tsinid = []
    role = []
    chapter = []
    squad = []
    snow_id = []
    tribe = []
    demandtype = []
    for i in range(int(total)):
        tsinid.append(request.form.get("tsinid"+str(i)))
        role.append(request.form.get("role"+str(i)))
        chapter.append(request.form.get("chapter"+str(i)))
        squad.append(request.form.get("squad"+str(i)))
        snow_id.append(request.form.get("snowid"+str(i)))
        tribe.append(request.form.get("tribe"+str(i)))
        demandtype.append(request.form.get("demand"+str(i)))

    result = roles.query.with_entities(roles.tsin_id, roles.snow_id).all()
    for i in result:
        if i[0].strip() in tsinid:
            return "TSIN ID already exists"
        if snow_id!=[]:
            if i[1].strip() in snow_id:
                print(i[1])
                return "SNOW ID already exists"
    for i in range(len(tsinid)):
        db.session.add(
            roles(
                tsin_id=tsinid[i],
                role=role[i],
                chapter=chapter[i],
                squad=squad[i],
                demand_type=demandtype[i],
                tribe=tribe[i],
                snow_id=snow_id[i],
                aspd_date=None,
                status="active",
                created_at=datetime.now(),
                modified_at=datetime.now(),
            )
        )
    db.session.commit()
    return total+" New Roles Created"


@app.route("/add-new-profile", methods=["GET", "POST"])
def newprofile():
    tsin = request.form["tsin_id"]
    phone = request.form["phone"]
    name = request.form["name"]
    email = request.form["email"]
    pan = request.form["pan"]
    company = request.form["company"]
    experience = request.form["experience"]
    location = request.form["location"]
    resume = request.files["resume"]
    db.session.add(
        candidates(
            tsin_id=tsin,
            candidate_name=name,
            pan=pan,
            candidate_email=email,
            current_stage="Profile Upload",
            status="active",
            request_raised_date=None,
            tsin_opened_date=None,
            resume_screened_date=None,
            l1_interview_date=None,
            l1_interviewer="",
            l2_interview_date=None,
            l2_interviewer="",
            l3_interview_date=None,
            l3_interviewer="",
            offer_rollout_date=None,
            joining_date=None,
            buddy_assignment_date=None,
            buddy_name="",
            candidate_dropout_date=None,
            candidate_dropout_reason="",
            resume="",
            resume_screened_remarks="",
            l1_interview_result="",
            l1_interview_remarks="",
            l2_interview_remarks="",
            l2_interview_result="",
            l3_interview_result="",
            l3_interview_remarks="",
            resume_remarks="",
            l1_completion=None, 
            l2_completion=None,
            l3_completion=None,
            phone=phone,
            current_location=location,
            current_company=company,
            experience=experience,
            candidate_joined_date=None,
            created_at=datetime.now(),
            modified_at=datetime.now(),
        )
    )
    db.session.commit()

    if resume:
        resume.save("static/resume")
        # minioClient.fput_object("resume", "resume" + str(name), "static/resume")
        # print(
        #     "resume is successfully uploaded as "
        #     "object 'resume" + str(name) + "' to bucket 'resume'."
        # )
    return redirect(url_for("main"))


@app.route("/apply-filter", methods=["GET", "POST"])
def apply_filter():
    tsin = request.form.get("val1")
    role = request.form.get("val2")
    chapter = request.form.get("val3")
    squad = request.form.get("val4")
    cname = request.form.get("val5")
    z = get_detailed_data()
    final = []
    for i in z:
        if tsin:
            if i["tsinid"].strip() == tsin.strip():
                if i not in final:
                    final.append(i)
        if role:
            if i["role"].strip() == role.strip():
                if i not in final:
                    final.append(i)
        if chapter:
            if i["chapter"].strip() == chapter.strip():
                if i not in final:
                    final.append(i)
        if squad:
            if i["squad"].strip() == squad.strip():
                if i not in final:
                    final.append(i)
        if cname:
            if i["candidate_name"].strip() == cname.strip():
                if i not in final:
                    final.append(i)
    return render_template("detailed.html", final=final)


@app.route("/edit-role", methods=["GET", "POST"])
def editrole():
    tsin = request.form.get("tsin_id")
    snow_id = request.form.get("snow_id")
    chapter = request.form.get("chapter")
    role = request.form.get("role")
    squad = request.form.get("squad")
    demandtype = request.form.get("demandtype")
    tribe = request.form.get("tribe")
    edit_check=request.form.get("editss")
    db.session.query(roles).filter(roles.tsin_id == tsin).update(
        {
            "snow_id": snow_id,
            "chapter": chapter,
            "role": role,
            "squad": squad,
            "demand_type": demandtype,
            "tribe": tribe,
            "status":"active",
            "modified_at": datetime.now(),
        }
    )
    db.session.commit()
    if edit_check=="onlyedit":
        return redirect(url_for("main"))
    return "Succesfully updated Role"


@app.route("/dashboard", methods=["GET", "POST"])
@token_forwarder
def dashboard(user):
    #print(token)
    # z, drop = total_offers()
    # new_pos = get_new_profiles()
    # weekly = 0
    # montly = 0
    # quarterly = 0
    # for i in z.values():
    #     for j in i:
    #         if j <= 14:
    #             weekly += 1
    #         if j <= 30:
    #             montly += 1
    #         if j <= 90:
    #             quarterly += 1
    # for i in z:
    #     z[i] = len(z[i])
    # offers = [weekly, montly, quarterly]
    # dropout_weekly = 0
    # dropout_montly = 0
    # dropout_quarterly = 0
    # for i in drop.values():
    #     for j in i:
    #         if j <= 14:
    #             dropout_weekly += 1
    #         if j <= 30:
    #             dropout_montly += 1
    #         if j <= 30:
    #             dropout_quarterly += 1
    # for i in drop:
    #     drop[i] = len(drop[i])
    # new_weekly = 0
    # new_monthly = 0
    # new_quarterly = 0
    # for i in new_pos.values():
    #     for j in i:
    #         if j <= 14:
    #             new_weekly += 1
    #         if j <= 30:
    #             new_monthly += 1
    #         if j <= 90:
    #             new_quarterly += 1
    roles, profileUpload, resume_screened, l1_interview, l2_interview, l3_interview, offer_rollout, buddy_assign, candidate_joined = get_role_wise_profiles()
    
    if(user):
        role=user.role
    else:
        role=None
    return render_template(
        "dashboard.html",
        role_wise_new=roles,
        profile_upload=profileUpload, 
        resume=resume_screened, 
        l1_interview=l1_interview,
        l2_interview=l2_interview,
        l3_interview= l3_interview,
        offer=offer_rollout,
        buddy=buddy_assign,
        candidate_joined=candidate_joined,
		user=role
    )
@app.route('/getData', methods = ['GET', 'POST'])
def getData():
    tsin_id = request.args.get('tsin_id')
    role = request.args.get('role')
    chapter = request.args.get('chapter')
    squad = request.args.get('squad')
    df=pd.read_sql_query("SELECT * FROM candidates JOIN roles on candidates.tsin_id=roles.tsin_id",engine)#db.engine)
    df=df.astype({'tsin_opened_date':'datetime64[ns]','resume_screened_date':'datetime64[ns]','l1_interview_date':'datetime64[ns]','l2_interview_date':'datetime64[ns]','l3_interview_date':'datetime64[ns]','offer_rollout_date':'datetime64[ns]','buddy_assignment_date':'datetime64[ns]','joining_date':'datetime64[ns]','candidate_dropout_date':'datetime64[ns]'})
    df['resume_sharing_time']=df['tsin_opened_date']-df['request_raised_date']
    df['l1_completion_time']=df['resume_screened_date']-df['l1_interview_date']
    df['l2_completion_time']=df['l2_interview_date']-df['l1_interview_date']
    df['l3_completion_time']=df['l3_interview_date']-df['l2_interview_date']
    df['time_to_offer']=df['offer_rollout_date']-df['request_raised_date']
    df['time_to_fill']=df['joining_date']-df['request_raised_date']
    
    if tsin_id=='all':
        if(role):
            df=df[df['role']==role]
        if(chapter):
            df=df[df['chapter']==chapter]
        if(squad):
            df=df[df['squad']==squad]
        if(df.empty):
            return {"KPIlabels":["Resume Selection Rate","L1 Selection Rate","L2 Selection Rate","L3 Selection Rate","Joining rate","Offer ratio"],
            "KPIdata":[],
            "SLAlabels":["Resume sharing","L1 completion","L2 completion","L3 completion","Buddy Assignment","Time to offer"],
            "SLAdata":[],
            "text":[[],[],[],[]]}
        t=df[df['offer_rollout_date'].notnull()]
        positives_kpi=[]
        negatives_kpi=[]
        kpilabels=["Resume Selection Rate","L1 Selection Rate","L2 Selection Rate","L3 Selection Rate","Joining rate","Offer ratio"]
        kpidata=np.array([df[df['resume_remarks']=='selected']['id'].count()/df['id'].count(),
        df[df['l1_interview_result']=='passed']['id'].count()/df[df['resume_remarks'].notnull()]['id'].count(),
        df[df['l2_interview_result']=='passed']['id'].count()/df[df['resume_remarks'].notnull()]['id'].count(),
        df[df['l3_interview_result']=='passed']['id'].count()/df[df['resume_remarks'].notnull()]['id'].count(),
        (t['id'].count()-t[t['candidate_dropout_date'].notnull()]['id'].count())/t['id'].count() if len(t)>0 else 0,
        (t['id'].count()-t[t['candidate_dropout_date'].notnull()]['id'].count())/df[df['l1_interview_result']=='passed']['id'].count() if len(t)>0 else 0])*100
        for i,j in enumerate(kpilabels):
            if kpidata[i]>50:
                positives_kpi.append(j+" increased upto {} %".format(kpidata[i]))
            else:
                negatives_kpi.append(j+" decreased to {} %".format(kpidata[i]))
        print(positives_kpi,negatives_kpi)

        slalabels=["Resume sharing","L1 completion","L2 completion","L3 completion","Buddy Assignment","Time to offer"]
        slametrics=[5,3,3,3,2,30]
        d1=np.array([df[df['resume_sharing_time']<pd.Timedelta('5 days')]['id'].count(),df[df['l1_completion_time']<pd.Timedelta('3 days')]['id'].count(), df[df['l2_completion_time']<pd.Timedelta('3 days')]['id'].count(),df[df['l3_completion_time']<pd.Timedelta('3 days')]['id'].count(),df[df['time_to_offer']<pd.Timedelta('3 days')]['id'].count(),df[df['time_to_fill']<pd.Timedelta('3 days')]['id'].count()])
        d2=np.array([df[df['resume_sharing_time']<pd.Timedelta('7 days')]['id'].count(),df[df['l1_completion_time']<pd.Timedelta('5 days')]['id'].count(), df[df['l2_completion_time']<pd.Timedelta('5 days')]['id'].count(),df[df['l3_completion_time']<pd.Timedelta('5 days')]['id'].count(),df[df['time_to_offer']<pd.Timedelta('5 days')]['id'].count(),df[df['time_to_fill']<pd.Timedelta('5 days')]['id'].count()]) - d1
        d3=np.array([df[df['resume_sharing_time']>pd.Timedelta('7 days')]['id'].count(),df[df['l1_completion_time']>pd.Timedelta('5 days')]['id'].count(), df[df['l2_completion_time']>pd.Timedelta('5 days')]['id'].count(),df[df['l3_completion_time']>pd.Timedelta('5 days')]['id'].count(),df[df['time_to_offer']>pd.Timedelta('5 days')]['id'].count(),df[df['time_to_fill']>pd.Timedelta('5 days')]['id'].count()])
        positives_sla=["{} resumes were cleared in {} process under {} days".format(d1[i],j,slametrics[i]) for i,j in enumerate(slalabels) if d1[i]>0]
        negatives_sla=["{} resumes took more than {} days in {} process".format(d3[i],slametrics[i],j) for i,j in enumerate(slalabels) if d3[i]>0]
        print(positives_sla,negatives_sla)
        return {"KPIlabels":kpilabels,
        "KPIdata":kpidata.tolist(),
        "SLAlabels":slalabels,
        "SLAdata":[d1.astype('int').tolist(),d2.astype('int').tolist(),d3.astype('int').tolist()],
        "text":[positives_kpi,negatives_kpi,positives_sla,negatives_sla]}
    else:
        df=df[df['tsin_id']==tsin_id]
        print(role,squad,chapter)
        if(role):
            df=df[df['roles']==role]
        if(chapter):
            df=df[df['chapter']==chapter]
        if(squad):
            df=df[df['squads']==squad]
        if(df.empty):
            return {"KPIlabels":["Resume Selection Rate","L1 Selection Rate","L2 Selection Rate","L3 Selection Rate","Joining rate","Offer ratio"],
            "KPIdata":[],
            "SLAlabels":["Resume sharing","L1 completion","L2 completion","L3 completion","Buddy Assignment","Time to offer"],
            "SLAdata":[],
            "text":[[],[],[],[]]}
        t=df[df['offer_rollout_date'].notnull()]
        #print(df[df['resume_screened_remarks']=='selected']['id'].count())
        kpilabels=["Resume Selection Rate","L1 Selection Rate","L2 Selection Rate","L3 Selection Rate","Joining rate","Offer ratio"]
        kpidata=np.array([df[df['resume_remarks']=='selected']['id'].count()/df['id'].count(),
        df[df['l1_interview_result']=='passed']['id'].count()/df[df['resume_remarks'].notnull()]['id'].count(),
        df[df['l2_interview_result']=='passed']['id'].count()/df[df['resume_remarks'].notnull()]['id'].count(),
        df[df['l3_interview_result']=='passed']['id'].count()/df[df['resume_remarks'].notnull()]['id'].count(),
        (t['id'].count()-t[t['candidate_dropout_date'].notnull()]['id'].count())/t['id'].count() if len(t)>0 else 0,
        (t['id'].count()-t[t['candidate_dropout_date'].notnull()]['id'].count())/df[df['l1_interview_result']=='passed']['id'].count() if len(t)>0 else 0])*100
        for i,j in enumerate(kpilabels):
            if kpidata[i]>50:
                positives_kpi.append(j+" increased upto {} %".format(kpidata[i]))
            else:
                negatives_kpi.append(j+" decreased to {} %".format(kpidata[i]))
        print(positives_kpi,negatives_kpi)

        slalabels=["Resume sharing","L1 completion","L2 completion","L3 completion","Buddy Assignment","Time to offer"]
        slametrics=[5,3,3,3,2,30]
        d1=np.array([df[df['resume_sharing_time']<pd.Timedelta('5 days')]['id'].count(),df[df['l1_completion_time']<pd.Timedelta('3 days')]['id'].count(), df[df['l2_completion_time']<pd.Timedelta('3 days')]['id'].count(),df[df['l3_completion_time']<pd.Timedelta('3 days')]['id'].count(),df[df['time_to_offer']<pd.Timedelta('3 days')]['id'].count(),df[df['time_to_fill']<pd.Timedelta('3 days')]['id'].count()])
        d2=np.array([df[df['resume_sharing_time']<pd.Timedelta('7 days')]['id'].count(),df[df['l1_completion_time']<pd.Timedelta('5 days')]['id'].count(), df[df['l2_completion_time']<pd.Timedelta('5 days')]['id'].count(),df[df['l3_completion_time']<pd.Timedelta('5 days')]['id'].count(),df[df['time_to_offer']<pd.Timedelta('5 days')]['id'].count(),df[df['time_to_fill']<pd.Timedelta('5 days')]['id'].count()]) - d1
        d3=np.array([df[df['resume_sharing_time']>pd.Timedelta('7 days')]['id'].count(),df[df['l1_completion_time']>pd.Timedelta('5 days')]['id'].count(), df[df['l2_completion_time']>pd.Timedelta('5 days')]['id'].count(),df[df['l3_completion_time']>pd.Timedelta('5 days')]['id'].count(),df[df['time_to_offer']>pd.Timedelta('5 days')]['id'].count(),df[df['time_to_fill']>pd.Timedelta('5 days')]['id'].count()])
        positives_sla=["{} resumes were cleared in {} process under {} days".format(d1[i],j,slametrics[i]) for i,j in enumerate(slalabels) if d1[i]>0]
        negatives_sla=["{} resumes took more than {} days in {} process".format(d3[i],slametrics[i],j) for i,j in enumerate(slalabels) if d3[i]>0]
        print(positives_sla,negatives_sla)

        return {"KPIlabels":kpilabels,
        "KPIdata":kpidata.tolist(),
        "SLAlabels":["Resume sharing","L1 completion","L2 completion","L3 completion","Buddy Assignment","Time to offer"],
        "SLAdata":[d1.astype('int').tolist(),d2.astype('int').tolist(),d3.astype('int').tolist()],
        "text":[positives_kpi,negatives_kpi,positives_sla,negatives_sla]}

@app.route('/add-chapter', methods = ['GET', 'POST'])
def addchapter():
    Chapter=request.form.get("chapter")
    chaps=Chapter.split(',')
    print(chaps)
    objects=[]
    for i in chaps:
        objects.append(chapter(chapter_name=i.strip()))
    db.session.bulk_save_objects(objects)
    db.session.commit()
    return "Chapter Added Succesfully"

@app.route('/add-role', methods = ['GET', 'POST'])
def addroler():
    Role=request.form.get("role")
    roless=Role.split(',')
    objects=[]
    for i in roless:
        objects.append(role(role_name=i.strip()))
    db.session.bulk_save_objects(objects)
    db.session.commit()
    return "Role Added Succesfully"

@app.route('/add-squad', methods = ['GET', 'POST'])
def addsquad():
    Squad=request.form.get("squad")
    squads=Squad.split(",")
    objects=[]
    for i in squads:
        objects.append(squad(squad_name=i.strip()))
    db.session.bulk_save_objects(objects)
    db.session.commit()
    return "Squad Added Succesfully"

@app.route('/add-tribe', methods = ['GET', 'POST'])
def addtribe():
    Tribe=request.form.get("tribe")
    tribes=Tribe.split(',')
    objects=[]
    for i in tribes:
        objects.append(tribe(tribe_name=i.strip()))
    db.session.bulk_save_objects(objects)
    db.session.commit()
    return "Tribe Added Succesfully"

@app.route('/add-program', methods = ['GET', 'POST'])
def addprogram():
    Program=request.form.get("program")
    programs=Program.split(',')
    objects=[]
    for i in programs:
        objects.append(program(program_name=i.strip()))
    db.session.bulk_save_objects(objects)
    db.session.commit()
    return "Program Added Succesfully"


@app.route('/add-project', methods = ['GET', 'POST'])
def addproject():
    Project=request.form.get("project")
    project_start=request.form.get("startdate")
    project_end=request.form.get("enddate")
    db.session.add(project(project_name=Project,start_date=project_start,end_date=project_end))
    db.session.commit()
    return "Project Added Succesfully"

@app.route('/login', methods = ['GET', 'POST'])
@token_forwarder
def login(user):
    if(user):
        role=user.role
        name=user.name
        email=user.email
    else:
        role=None
        name=None
        email=None
    return render_template("login.html", user=role, name=name, email=email)

@app.route('/signup', methods = ['GET', 'POST'])
@token_forwarder
def signup(user):
    if(user):
        role=user.role
    else:
        role=None
    users = Users.query.all()
    return render_template("signup.html", users=users, user=role)

@app.route('/chapter', methods = ['GET', 'POST'])
@token_forwarder
def addChapter(user):
    if(user):
        role=user.role
    else:
        role=None
    return render_template("add-chapter.html", user=role)

@app.route('/role', methods = ['GET', 'POST'])
@token_forwarder
def addRole(user):
    if(user):
        role=user.role
    else:
        role=None
    return render_template("add-role.html", user=role)
    
@app.route('/squad', methods = ['GET', 'POST'])
@token_forwarder
def addSquad(user):
    if(user):
        role=user.role
    else:
        role=None
    return render_template("add-squad.html", user=role)

@app.route('/tribe', methods = ['GET', 'POST'])
@token_forwarder
def addTribe(user):
    if(user):
        role=user.role
    else:
        role=None
    return render_template("add-tribe.html", user=role)

@app.route('/program', methods = ['GET', 'POST'])
@token_forwarder
def addProgram(user):
    if(user):
        role=user.role
    else:
        role=None
    return render_template("add-program.html", user=role)

@app.route('/project', methods = ['GET', 'POST'])
@token_forwarder
def addProject(user):
    if(user):
        role=user.role
    else:
        role=None
    return render_template("add-project.html", user=role)

@app.route('/hired-candidates', methods = ['GET', 'POST'])
@token_forwarder
def cands(user):
    if(user):
        role=user.role
    else:
        role=None
    ename = request.form.get("val1")
    erole = request.form.get("val2")
    eprogram = request.form.get("val3")
    eproject = request.form.get("val4")
    roles= Employee.query.with_entities(Employee.role).distinct(Employee.role).all()
    unique_roles=[i[0] for i in roles]
    names= Employee.query.with_entities(Employee.name).distinct(Employee.name).all()
    unique_names=[i[0] for i in names]
    programs=program.query.with_entities(program.program_name).distinct(program.program_name).all()
    unique_programs=[i[0] for i in programs]
    projects=project.query.with_entities(project.project_name).distinct(project.project_name).all()
    unique_projects=[i[0] for i in projects]
    final=[]
    current_year=datetime.now().year
    # alloc=Allocations.query.join(project, Employee).filter(Allocations.year>=current_year, Allocations.year<=current_year+1).all()
    alloc=db.session.query(Allocations, project, Employee).filter(Allocations.emp_id == Employee.id,Allocations.project_id==project.id, Allocations.year>=current_year, Allocations.year<=current_year+1).all()
    print(current_year)
    user_wise={}
    # for i in unique_names:
    #     user_wise[i]={current_year:{1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0},current_year+1:{1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0}}
    print(user_wise)
    for a,p,e in alloc:
        user_wise[e.name+ p.project_name]={current_year:{1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0},current_year+1:{1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0}}
    for a,p,e in alloc:
        if a.status=='inactive':
            continue
        # project_name=project.query.with_entities(project.project_name).filter_by(id=i.project_id).first()
        program_name=program.query.with_entities(program.program_name).filter_by(id=a.program_id).first()
        # user=Employee.query.filter_by(id=i.emp_id).first()
        user_wise[e.name+ p.project_name][a.year][a.month]=a.allocation_percentage
        user_l=[e.name, e.role, program_name[0], p.project_name]
        if user_l not in final:
            final.append(user_l)
    for i in final:
        i.append(user_wise[i[0]+i[3]])
    secfinal=[]
    if ename:
        secf=[j for j in final if j[0]==ename]
        secfinal+=secf
    if erole:
        secf=[j for j in final if j[1]==erole]
        secfinal+=secf
    if eprogram:
        print('yes')
        secf=[j for j in final if j[2]==eprogram]
        secfinal+=secf
    if eproject:
        secf=[j for j in final if j[3]==eproject]
        secfinal+=secf
    print(secfinal)
    if ename or erole or eprogram or eproject:
        return render_template("details_emp.html", user=role, emp=secfinal, names=unique_names, roles=unique_roles, programs=unique_programs, projects=unique_projects, year=current_year)
    return render_template("hired_employees.html", user=role, emp=final, names=unique_names, roles=unique_roles, programs=unique_programs, projects=unique_projects, year=current_year)


@app.route("/upload-resource", methods=["GET", "POST"])
def uploadDr():
    f = request.files["file"]  # File input
    if not f:
        return "No file attached"

    global filename
    filename = f.filename  # changing global value of filename

    path = "{}/{}".format("static", filename)
    f.save(path)
    x = path.split(".")[-1]

    # reading filedata start
    if x == "xlsx":
        new_wb = load_workbook(path)
        Dataframe = pd.read_excel(new_wb, engine="openpyxl")
    elif x == "csv":
        Dataframe = pd.read_csv(path, encoding="ISO-8859-1")
    else:
        return "Please upload the file in xlsx or csv only"
    # reading filedata end
    dict_of_records = Dataframe.to_dict(orient="records")
    print(Dataframe.columns)
    for i in dict_of_records:
        id=i['Employee code']
        name=i['Employee Name']
        role=i['Role']
        programm=i['Program']
        projectt=i['Project/Use case']
        if i['Employee code']==0 or pd.isna(projectt) or pd.isna(i['Employee code']):
            continue
        names=Employee.query.with_entities(Employee.id).all()
        exist=False
        for j in names:
            if str(id)==str(j[0]):
                print('yes')
                exist=True
        if exist==False:
            db.session.add(Employee(id=id, name = name, role=role))
            db.session.commit()
        
        year_wise=[i['2023 Jan'],i['2023 Feb'],i['2023 Mar'],i['2023 April'],i['2023 May'],i['2023 June'],i['2023 Jul'],i['2023 Aug'],i['2023 Sep'],i['2023 Oct'],i['2023 Nov'],i['2023 Dec']]
        x = float("nan")
        start=None
        end=None
        year=2023
        project_id=project.query.with_entities(project.id).filter_by(project_name=str(projectt)).first()
        program_id=program.query.with_entities(program.id).filter_by(program_name=str(programm)).first()
        squad_id=squad.query.with_entities(squad.id).filter_by(squad_name=str(programm)).first()
        if project_id:
            project_id=project_id[0]
        if program_id:
            program_id=program_id[0]
        if squad_id:
            squad_id=squad_id[0]
        for j in range(len(year_wise)):
            if pd.isna(year_wise[j]) or isinstance(year_wise[j], str):
                continue
            else:
                month=int(j)+1
                if year_wise[j]=='=30%*0+100%':
                    percentage_alloc=100
                elif year_wise[j]=='=70%*0' or year_wise[j]=='=25%*0':
                    percentage_alloc=0
                elif pd.isna(year_wise[j]):
                    percentage_alloc=0
                else:
                    percentage_alloc=year_wise[j]*100
                db.session.add(Allocations(emp_id=id,project_id=project_id, program_id=program_id,squad_id=squad_id, month=month,year=year,allocation_percentage=percentage_alloc, status="active"))
                db.session.commit()
        
        # db.session.flush()
        print(project_id,squad_id, program_id)
    print("added to db")
    return "ok"


@app.route('/edit-allocation', methods = ['GET', 'POST'])
def update_allocations():
    empdetails=request.form.get('emp_details')
    jan1=request.form.get('jan1')
    feb1=request.form.get('feb1')
    mar1=request.form.get('mar1')
    apr1=request.form.get('apr1')
    may1=request.form.get('may1')
    jun1=request.form.get('jun1')
    jul1=request.form.get('jul1')
    aug1=request.form.get('aug1')
    sep1=request.form.get('sep1')
    oct1=request.form.get('oct1')
    nov1=request.form.get('nov1')
    dec1=request.form.get('dec1')
    jan2=request.form.get('jan2')
    feb2=request.form.get('feb2')
    mar2=request.form.get('mar2')
    apr2=request.form.get('apr2')
    may2=request.form.get('may2')
    jun2=request.form.get('jun2')
    jul2=request.form.get('jul2')
    aug2=request.form.get('aug2')
    sep2=request.form.get('sep2')
    oct2=request.form.get('oct2')
    nov2=request.form.get('nov2')
    dec2=request.form.get('dec2')
    first=[jan1, feb1, mar1, apr1, may1, jun1, jul1, aug1, sep1, oct1, nov1, dec1]
    second=[jan2, feb2, mar2, apr2, may2, jun2, jul2, aug2, sep2, oct2, nov2, dec2]
    year=[datetime.now().year, datetime.now().year+1]
    empname, project_name=empdetails.split('_')
    empid=Employee.query.with_entities(Employee.id).filter_by(name=empname).first()
    project_id=project.query.with_entities(project.id).filter_by(project_name=project_name).first()
    print(empid,project_id)
    for i in range(2):
        for j in range(1,13):
            if i==0:
                check=db.session.query(Allocations).filter(
                            Allocations.emp_id==empid[0],
                            Allocations.project_id==project_id[0],
                            Allocations.month==j,
                            Allocations.year==year[i]
                        ).first()
                print(check)
                if check:
                    db.session.query(Allocations).filter(
                                Allocations.emp_id==empid[0],
                                Allocations.project_id==project_id[0],
                                Allocations.month==j,
                                Allocations.year==year[i]
                            ).update(
                                {
                                    "allocation_percentage":first[j-1]
                                }
                    )
                else:
                    program=Allocations.query.with_entities(Allocations.program_id, Allocations.squad_id).filter_by(
                        emp_id=empid[0],
                        project_id=project_id[0]
                    ).first()
                    print(program)
                    db.session.add(Allocations(emp_id=empid[0],program_id=program[0], squad_id=program[1], project_id=project_id[0], allocation_percentage= first[j-1], month=j, year=year[i], status="active" ))
                db.session.commit()
            else:
                check=db.session.query(Allocations).filter(
                            Allocations.emp_id==empid[0],
                            Allocations.project_id==project_id[0],
                            Allocations.month==j,
                            Allocations.year==year[i]
                        ).first()
                print(check)
                if check:
                    db.session.query(Allocations).filter(
                                Allocations.emp_id==empid[0],
                                Allocations.project_id==project_id[0],
                                Allocations.month==j,
                                Allocations.year==year[i]
                            ).update(
                                {
                                    "allocation_percentage":second[j-1]
                                }
                    )
                else:
                    program=Allocations.query.with_entities(Allocations.program_id,  Allocations.squad_id).filter_by(
                        emp_id=empid[0],
                        project_id=project_id[0]
                    ).first()
                    print(program)
                    db.session.add(Allocations(emp_id=empid[0],program_id=program[0], squad_id=program[1], project_id=project_id[0], allocation_percentage= second[j-1], month=j, year=year[i], status="active" ))
                db.session.commit()
    return "done"


@app.route('/add-new-allocation', methods = ['GET', 'POST'])
def add_allocations():
    ename=request.form.get("name")
    erole=request.form.get("role")
    eprogram=request.form.get("program")
    eproject=request.form.get("project")
    jan1=request.form.get('jan1')
    feb1=request.form.get('feb1')
    mar1=request.form.get('mar1')
    apr1=request.form.get('apr1')
    may1=request.form.get('may1')
    jun1=request.form.get('jun1')
    jul1=request.form.get('jul1')
    aug1=request.form.get('aug1')
    sep1=request.form.get('sep1')
    oct1=request.form.get('oct1')
    nov1=request.form.get('nov1')
    dec1=request.form.get('dec1')
    jan2=request.form.get('jan2')
    feb2=request.form.get('feb2')
    mar2=request.form.get('mar2')
    apr2=request.form.get('apr2')
    may2=request.form.get('may2')
    jun2=request.form.get('jun2')
    jul2=request.form.get('jul2')
    aug2=request.form.get('aug2')
    sep2=request.form.get('sep2')
    oct2=request.form.get('oct2')
    nov2=request.form.get('nov2')
    dec2=request.form.get('dec2')
    first=[jan1, feb1, mar1, apr1, may1, jun1, jul1, aug1, sep1, oct1, nov1, dec1]
    second=[jan2, feb2, mar2, apr2, may2, jun2, jul2, aug2, sep2, oct2, nov2, dec2]
    year=[datetime.now().year, datetime.now().year+1]
    names=Employee.query.with_entities(Employee.name).all()
    exist=False
    for j in names:
        if str(ename)==str(j[0]):
            print('yes')
            exist=True
    if exist==False:
        db.session.add(Employee(id="2020002181", name = ename, role=erole))
        db.session.commit()
    eid=Employee.query.with_entities(Employee.id).filter_by(name=ename).first()
    program_id=program.query.with_entities(program.id).filter_by(program_name=eprogram).first()
    project_id=project.query.with_entities(project.id).filter_by(project_name=eproject).first()
    squad_id=squad.query.with_entities(squad.id).filter_by(squad_name=eproject).first()
    if not squad_id:
        squad_id=None
    else:
        squad_id=squad_id[0]
    for i in range(2):
        for j in range(1,13):
            if i==0:
                db.session.add(Allocations(emp_id=eid[0], program_id=program_id[0],project_id=project_id[0],squad_id=squad_id,status="active", month=j, year=year[i],allocation_percentage=first[j-1]))
            else:
                db.session.add(Allocations(emp_id=eid[0], program_id=program_id[0],project_id=project_id[0],squad_id=squad_id,status="active", month=j, year=year[i],allocation_percentage=second[j-1]))
            db.session.commit()
    return "New Allocation Added"

@app.route('/billing', methods = ['GET', 'POST'])
def generate_billing():
    f = request.files["file"]  # File input
    if not f:
        return "No file attached"

    global filename
    filename = f.filename  # changing global value of filename

    path = "{}/{}".format("static", filename)
    f.save(path)
    x = path.split(".")[-1]

    # reading filedata start
    if x == "xlsx":
        new_wb = load_workbook(path)
        Dataframe = pd.read_excel(new_wb, engine="openpyxl")
    elif x == "csv":
        Dataframe = pd.read_csv(path, encoding="ISO-8859-1")
    else:
        return "Please upload the file in xlsx or csv only"
    # reading filedata end
    dict_of_records = Dataframe.to_dict(orient="records")
    print(Dataframe.columns)
    return "done"

if __name__ == "__main__":
    app.run()
